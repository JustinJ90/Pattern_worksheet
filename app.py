#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pattern Worksheet Generator - Final Version
- Inputs: Name & Grade (Optional)
- Preview Mode: Opens in new tab instead of auto-download
- Features: Random mixing, Teacher's Guide(Page 2), A4 Layout
"""

from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import os
import glob
import random
from datetime import datetime

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
DB_FOLDER = os.path.join(BASE_DIR, 'databases')

os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(DB_FOLDER, exist_ok=True)

def setup_korean_font():
    try:
        local_font = os.path.join(BASE_DIR, 'fonts', 'NanumGothic.ttf')
        if os.path.exists(local_font):
            pdfmetrics.registerFont(TTFont('KoreanFont', local_font))
            return 'KoreanFont'
        return 'Helvetica' 
    except:
        return 'Helvetica'

KOREAN_FONT = setup_korean_font()

def load_patterns_from_excel(filename):
    file_path = os.path.join(DB_FOLDER, filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"DB 파일을 찾을 수 없습니다: {filename}")

    wb = openpyxl.load_workbook(file_path)
    
    ws_overview = wb["Pattern Overview"]
    pattern_info = {}
    for row in ws_overview.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            pattern_info[int(row[0])] = {
                'number': int(row[0]),
                'name': str(row[1]),
                'unit': str(row[3]) if len(row) > 3 and row[3] else 'Level A'
            }
            
    ws_detail = wb["Pattern Details"]
    patterns = {}
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        try:
            p_num = int(row[0])
            section = row[2]
            content = row[4]
            answer = row[5] if len(row) > 5 and row[5] else ""
            
            if p_num not in patterns:
                patterns[p_num] = {
                    'pattern_num': p_num,
                    'pattern_name': pattern_info.get(p_num, {}).get('name', ''),
                    'unit': pattern_info.get(p_num, {}).get('unit', 'Level A'),
                    'speaking1': [], 'speaking2': [], 'unscramble': []
                }
            
            if section == 'Speaking I':
                patterns[p_num]['speaking1'].append(content)
            elif section == 'Speaking II':
                patterns[p_num]['speaking2'].append((content, answer))
            elif section == 'Unscramble':
                scrambled = row[6].strip('()') if row[6] else ""
                patterns[p_num]['unscramble'].append((content, scrambled, answer))
        except:
            continue
            
    return patterns

def distribute_questions(selected_patterns, target_count=5):
    result = {'speaking1': [], 'speaking2': [], 'unscramble': []}
    if not selected_patterns: return result
    
    pattern_count = len(selected_patterns)
    items_per = target_count // pattern_count
    remainder = target_count % pattern_count
    
    for section in ['speaking1', 'speaking2', 'unscramble']:
        for i, p in enumerate(selected_patterns):
            count = items_per + (1 if i < remainder else 0)
            pool = p[section][:]
            random.shuffle(pool)
            result[section].extend(pool[:count])
            
    return result

# --- PDF 생성 (이름/학년 파라미터 추가) ---
def create_worksheet(pattern_data, selected_patterns, output_path, book_title, student_name="", student_grade=""):
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=10*mm,
        bottomMargin=10*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    story = []
    
    p_nums = ", ".join([str(p['pattern_num']) for p in selected_patterns])
    clean_book_title = book_title.replace('.xlsx', '')
    
    # Styles
    title_style = ParagraphStyle('Title', fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=5)
    section_style = ParagraphStyle('Section', fontSize=11, fontName='Helvetica-Bold', spaceBefore=0, spaceAfter=0)
    item_style = ParagraphStyle('Item', fontSize=10, fontName='Helvetica', leftIndent=0, spaceBefore=2, spaceAfter=2)
    item_kr_style = ParagraphStyle('ItemKr', fontSize=10, fontName=KOREAN_FONT, leftIndent=0, spaceBefore=2, spaceAfter=2)
    line_style = ParagraphStyle('Line', fontSize=10, fontName='Helvetica', spaceAfter=0)
    
    # === PAGE 1: Student Worksheet ===
    
    story.append(Paragraph("<b>Weekly Test</b>", title_style))
    story.append(Paragraph(f"<b>{clean_book_title} - Patterns: {p_nums}</b>", title_style))
    
    # [수정됨] 이름/날짜/학년 표시 로직
    # 이름이 입력되면 그 이름을 넣고, 아니면 밑줄을 넣음
    display_name = f"NAME: {student_name}" if student_name else "NAME: _______________________________"
    
    name_date_data = [[
        Paragraph(display_name, ParagraphStyle('Name', fontSize=12, fontName=KOREAN_FONT)), # 한글 이름 지원을 위해 폰트 변경
        Paragraph("DATE: _____ / _____", ParagraphStyle('Date', fontSize=12, fontName='Helvetica', alignment=TA_RIGHT))
    ]]
    name_date_table = Table(name_date_data, colWidths=[120*mm, 50*mm])
    name_date_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    story.append(name_date_table)
    story.append(Spacer(1, 4*mm))
    
    # Speaking I
    story.append(Paragraph("<b>◈ Speaking I - Answer the questions</b>", section_style))
    story.append(Spacer(1, 2*mm))
    for idx, question in enumerate(pattern_data['speaking1'][:5], 1):
        story.append(Paragraph(f"{idx}. {question}", item_style))
    story.append(Spacer(1, 4*mm))
    
    # Speaking II
    story.append(Paragraph("<b>◈ Speaking II - Say in English</b>", section_style))
    story.append(Spacer(1, 2*mm))
    for idx, (korean, answer) in enumerate(pattern_data['speaking2'][:5], 1):
        story.append(Paragraph(f"{idx}. {korean}", item_kr_style))
    story.append(Spacer(1, 4*mm))
    
    # Speaking III
    story.append(Paragraph("<b>◈ Speaking III - With your teacher</b>", section_style))
    story.append(Spacer(1, 2*mm))
    for idx in range(1, 6):
        story.append(Paragraph(f"{idx}. Pattern {idx}", item_style))
    story.append(Spacer(1, 4*mm))
    
    # Unscramble
    story.append(Paragraph("<b>◈ Unscramble</b>", section_style))
    story.append(Spacer(1, 2*mm))
    for idx, (korean, words, answer) in enumerate(pattern_data['unscramble'][:5], 1):
        story.append(Paragraph(f"{idx}. {korean} ({words})", item_kr_style))
        story.append(Spacer(1, 7*mm)) 
        story.append(Paragraph("_" * 85, line_style))
        story.append(Spacer(1, 3*mm))
    
    # Footer [수정됨] 학년(GRADE) 표시
    story.append(Spacer(1, 5*mm))
    
    display_grade = f"<b>GRADE: {student_grade}</b>" if student_grade else "<b>GRADE:</b>"
    
    footer_data = [[
        Paragraph(display_grade, ParagraphStyle('Footer', fontSize=12, fontName=KOREAN_FONT)), # 한글 학년 지원
        "",
        Paragraph("<b>REMARK:</b>", ParagraphStyle('Footer', fontSize=12, fontName='Helvetica-Bold'))
    ]]
    footer_table = Table(footer_data, colWidths=[40*mm, 40*mm, 90*mm])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'LEFT'),
    ]))
    story.append(footer_table)
    
    # === PAGE 2: Teacher's Guide ===
    story.append(PageBreak())
    
    story.append(Paragraph("<b>Teacher's Guide (Answer Key)</b>", title_style))
    story.append(Paragraph(f"<b>{clean_book_title} - Patterns: {p_nums}</b>", title_style))
    story.append(Spacer(1, 10*mm))
    
    story.append(Paragraph("<b>◈ Speaking II Answers</b>", section_style))
    story.append(Spacer(1, 3*mm))
    for idx, (korean, answer) in enumerate(pattern_data['speaking2'][:5], 1):
        story.append(Paragraph(f"<b>{idx}.</b> {answer}", item_style))
    
    story.append(Spacer(1, 10*mm))
    
    story.append(Paragraph("<b>◈ Unscramble Answers</b>", section_style))
    story.append(Spacer(1, 3*mm))
    for idx, (korean, words, answer) in enumerate(pattern_data['unscramble'][:5], 1):
        story.append(Paragraph(f"<b>{idx}.</b> {answer}", item_style))

    doc.build(story)

# --- 라우트 설정 ---

@app.route('/')
def index():
    files = glob.glob(os.path.join(DB_FOLDER, "*.xlsx"))
    books = sorted([os.path.basename(f) for f in files])
    return render_template('index.html', books=books)

@app.route('/get_patterns/<filename>')
def get_patterns(filename):
    try:
        patterns = load_patterns_from_excel(filename)
        pattern_list = []
        for p_num in sorted(patterns.keys()):
            pattern_list.append({
                'number': p_num,
                'name': patterns[p_num]['pattern_name'],
                'unit': patterns[p_num]['unit']
            })
        return jsonify({'success': True, 'patterns': pattern_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        book_filename = data.get('book')
        selected_nums = data.get('patterns', [])
        
        # [수정됨] 클라이언트에서 보낸 이름과 학년 받기
        student_name = data.get('name', '')
        student_grade = data.get('grade', '')
        
        if not book_filename or not selected_nums:
            return jsonify({'error': 'Book or Patterns missing'}), 400
            
        all_patterns = load_patterns_from_excel(book_filename)
        selected_data = []
        for num in selected_nums:
            if int(num) in all_patterns:
                selected_data.append(all_patterns[int(num)])
                
        final_questions = distribute_questions(selected_data)
        
        filename = f"Worksheet_{datetime.now().strftime('%m%d_%H%M%S')}.pdf"
        output_path = os.path.join(OUTPUT_FOLDER, filename)
        
        # [수정됨] PDF 생성 함수에 이름/학년 전달
        create_worksheet(final_questions, selected_data, output_path, book_filename, student_name, student_grade)
        
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000, debug=True)